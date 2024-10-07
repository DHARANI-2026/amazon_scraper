from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import time
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as ExcelImage
from selenium import webdriver
import os

# Create a folder to store images
if not os.path.exists('images'):
    os.makedirs('images')

# Configure Chrome to ignore certificate errors
chrome_options = Options()
chrome_options.add_argument('--ignore-certificate-errors')

# Initialize WebDriver using WebDriver Manager (auto-download ChromeDriver)
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

# Open Amazon search results page for chocolates
driver.get("https://www.amazon.in/s?k=chocolates")

# Wait for the page to load completely
time.sleep(5)

# Create an empty list to store the product details
products = []

# Scrape product details (image, name, price, rating, reviews, availability)
for i in range(6):  # Change the range for more products
    try:
        # Get product image URL and download the image
        try:
            image_url = driver.find_elements(By.CSS_SELECTOR, ".s-image")[i].get_attribute("src")
            image_response = requests.get(image_url)
            image_path = f'images/product_{i}.png'
            with open(image_path, 'wb') as file:
                file.write(image_response.content)
        except:
            image_path = None
        
        # Get product name (updated selector)
        try:
            product_name_elements = driver.find_elements(By.CSS_SELECTOR, "span.a-size-base-plus.a-color-base.a-text-normal")
            if len(product_name_elements) > i:
                product_name = product_name_elements[i].text
            else:
                product_name = "Name not available"
        except:
            product_name = "Name not available"
        
        # Get product price
        try:
            product_price_elements = driver.find_elements(By.CSS_SELECTOR, ".a-price-whole")
            if len(product_price_elements) > i:
                product_price = product_price_elements[i].text
            else:
                product_price = "Price not available"
        except:
            product_price = "Price not available"
        
        # Get product rating (updated CSS selector)
        try:
            product_rating_elements = driver.find_elements(By.CSS_SELECTOR, ".a-icon-alt")
            if len(product_rating_elements) > i:
                product_rating = product_rating_elements[i].get_attribute("innerHTML").split()[0]
            else:
                product_rating = "Rating not available"
        except:
            product_rating = "Rating not available"
        
        # Get product review count (updated CSS selector)
        try:
            product_reviews_elements = driver.find_elements(By.CSS_SELECTOR, ".s-link-style .s-underline-text")
            if len(product_reviews_elements) > i:
                product_reviews = product_reviews_elements[i].text
            else:
                product_reviews = "Reviews not available"
        except:
            product_reviews = "Reviews not available"

        # Get product availability (updated selector)
        try:
            product_availability_elements = driver.find_elements(By.CSS_SELECTOR, ".a-size-base.a-color-secondary")
            if len(product_availability_elements) > i:
                product_availability = product_availability_elements[i].text
            else:
                product_availability = "Availability not available"
        except:
            product_availability = "Availability not available"
        
        # Append the scraped data to the list (with image path)
        products.append([product_name, product_price, product_rating, product_reviews, product_availability, image_path])

    except Exception as e:
        print(f"Error: {e}")
        continue

# Close the browser
driver.quit()

# Save product details to Excel
wb = Workbook()
ws = wb.active
ws.title = "Amazon Chocolates"

# Add headers to the Excel sheet
headers = ["Product Name", "Product Price", "Product Rating", "Product Reviews", "Product Availability", "Product Image"]
ws.append(headers)

# Set bold font for headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Insert product data into Excel and add images
for row_num, product in enumerate(products, 2):
    ws.append([product[0], product[1], product[2], product[3], product[4]])

    # Insert image in the 6th column
    if product[5]:  # If image path exists
        img = ExcelImage(product[5])
        img.width, img.height = 100, 100  # Dynamically change image size based on its dimensions
        ws.row_dimensions[row_num].height = 75  # Adjust row height for images
        ws.add_image(img, f'F{row_num}')  # Add image to the cell

# Dynamically adjust column widths based on image size and content
for col in range(1, ws.max_column + 1):
    column_letter = get_column_letter(col)
    if col == 6:  # For image column
        ws.column_dimensions[column_letter].width = 18  # Adjust width for images
    else:
        ws.column_dimensions[column_letter].width = 25  # Set a default width for text columns

# Apply center alignment to all cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Save the Excel workbook
wb.save("amazon_chocolates_with_images.xlsx")

print("Data scraped and saved to 'amazon_chocolates_with_images.xlsx' successfully!")
